#!/usr/bin/env python3
"""\
Generate an .xlsx table of the Qx supported SDI generator standards
via the REST API

Usage: 
    generate_standards_sheet.py [--output=<filename>] HOSTNAME
    generate_standards_sheet.py --version
    generate_standards_sheet.py --help    

Generate an Excel spreadsheet containing the complete list of generator 
standards that a Qx reports as supporting.

Arguments:
    HOSTNAME                Hostname of the Qx (e.g. qx-020123.phabrix.local)

Options:
    --output=<filename>     Rest API port on the target Qx  [default: standards.xlsx]
    -h, --help              Usage help
    --version               Show version and exit

"""

from autolib.factory import make_qx
from autolib.models.qxseries.operationmode import OperationMode
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle, Font
from docopt import docopt


def generate(arguments):
    """\
    Connect to the specified Qx and query the standards list and then generate
    a simple Excel spreadsheet.
    """

    with make_qx(arguments["HOSTNAME"]) as qx:

        standards = qx.generator.get_standards()

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"{type(qx).__name__} SDI Generator Standards"

        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=True)
        highlight.fill = PatternFill("solid", fgColor="C3CCDB")

        worksheet.cell(row=1, column=1, value='SDI Data Rate').style = highlight
        worksheet.cell(row=1, column=2, value='Resolution / Frame Rate').style = highlight
        worksheet.cell(row=1, column=3, value='Color Mapping').style = highlight
        worksheet.cell(row=1, column=4, value='Format / Gamut').style = highlight

        row = 2
        worksheet.column_dimensions[get_column_letter(1)].width = 30.0
        worksheet.column_dimensions[get_column_letter(2)].width = 30.0
        worksheet.column_dimensions[get_column_letter(3)].width = 30.0
        worksheet.column_dimensions[get_column_letter(4)].width = 40.0

        for sdi_rate, sdi_standards in standards.items():
            for aspect, colours in sdi_standards.items():
                for colour, formats in colours.items():
                    for format_gamut in formats:
                        # Add a cell in column 1 for the resolution
                        worksheet.cell(row=row, column=1, value=f"{str(sdi_rate)} Gbps")
                        # Add a cell in column 1 for the resolution
                        worksheet.cell(row=row, column=2, value=aspect)
                        # Add a cell in column 2 for the colour
                        worksheet.cell(row=row, column=3, value=colour)
                        # Add a cell in column 3 for the format
                        worksheet.cell(row=row, column=4, value=format_gamut)
                        row += 1

        worksheet.auto_filter.ref = 'A1:D1'
        workbook.save(arguments.get("--output"))


if __name__ == "__main__":
    docopt_arguments = docopt(__doc__, version='0.0.1')
    generate(docopt_arguments)
